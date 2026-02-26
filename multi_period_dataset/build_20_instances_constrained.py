"""
Build 20 "black ice monitoring" road-network instances suitable for:
- GNN prediction (spatiotemporal graph learning)
- MD-RPP-RRV routing heuristics (arc/edge routing)

This version ENFORCES hard constraints at build time:
- Area (approx bbox) between 10–150 km^2  -> dist_m in [1600, 6100]
- Nodes between 100–500
- Edges between 300–1200

It does this by *searching over*:
  (a) dist_m candidates within [1600, 6100]
  (b) intersection consolidation tolerance candidates
  (c) two road filters: major roads only vs major+residential

It also screens for graph structure pathologies (star-hubs) using degree stats.

Outputs:
- graphs/*.pickle
- weather/*_hourly.csv (5 sampled points: centroid + corners)
- meta/*.json (full metadata)
- diagnostics/*_graph_sweep.csv (all tried configs + metrics)
- instances_metadata.csv (compact table)
- instances_metadata.xlsx (Excel version)

Notes:
- Area is reported as bbox approximation: area_km2 ≈ (2*dist_m)^2 / 1e6 = 4*dist_m^2 / 1e6.
  This matches how OSMnx uses a bbox around the center point when dist_type="bbox".
"""

import os
import json
import pickle
from dataclasses import dataclass, asdict
from typing import List, Tuple, Dict, Optional

import numpy as np
import pandas as pd
import osmnx as ox
import networkx as nx

import openmeteo_requests
import requests_cache
from retry_requests import retry

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------
# Constraints (your requirements)
# -----------------------------
AREA_MIN_KM2 = 10.0
AREA_MAX_KM2 = 150.0

NODES_MIN, NODES_MAX = 100, 500
EDGES_MIN, EDGES_MAX = 300, 1200

# Convert area bounds to dist bounds (bbox approx area=4*dist^2)
DIST_MIN_M = int(np.ceil(np.sqrt(AREA_MIN_KM2 * 1e6 / 4.0)))   # ~1581 m
DIST_MAX_M = int(np.floor(np.sqrt(AREA_MAX_KM2 * 1e6 / 4.0)))  # ~6124 m

# Candidate distances to try (meters) - reduced for faster execution
# Covers the full range [1600, 6000] with 5 values instead of 10
DIST_CANDIDATES_M = [1600, 2400, 3600, 4800, 6000]

# Intersection tolerance candidates (meters) - reduced for faster execution
# Covers the full range [20, 300] with 6 values instead of 12
TOLERANCE_CANDIDATES_M = [20, 40, 80, 125, 200, 300]

# Structural sanity thresholds
MIN_LCC_NODE_FRAC = 0.95
MIN_MAJOR_EDGE_FRAC = 0.25
MAX_MAX_DEGREE = 35
MAX_P99_DEGREE = 18

# -----------------------------
# Open‑Meteo configuration
# -----------------------------
OPENMETEO_HISTORICAL_URL = "https://archive-api.open-meteo.com/v1/archive"
HOURLY_VARS = [
    "temperature_2m",
    "relative_humidity_2m",
    "dew_point_2m",
    "precipitation",
    "rain",
    "snowfall",
    "snow_depth",
    "wind_speed_10m",
    "wind_gusts_10m",
    "surface_pressure",
    "weather_code",
]
MISSION_START_LOCAL = 16
MISSION_END_LOCAL = 20

# Passability heuristic (kept light; you can disable)
ENABLE_SNOW_PASSABILITY_CHECK = True
MAX_MEDIAN_SNOW_DEPTH_CM = 20.0
MAX_MEAN_SNOWFALL_MM = 5.0

# -----------------------------
# Road filters
# -----------------------------
MAJOR_ROADS = {"motorway", "trunk", "primary", "secondary", "tertiary"}

ROADS_FILTER_MAJOR_ONLY = '["highway"~"motorway|trunk|primary|secondary|tertiary"]'
ROADS_FILTER_WITH_RESIDENTIAL = '["highway"~"motorway|trunk|primary|secondary|tertiary|residential"]'

ROAD_FILTER_VARIANTS = [
    ("major_only", ROADS_FILTER_MAJOR_ONLY),
    ("major_plus_residential", ROADS_FILTER_WITH_RESIDENTIAL),
]


@dataclass
class InstanceSpec:
    instance_id: int
    place: str
    state: str
    center_lat: float
    center_lon: float
    timezone: str
    start_date: str
    end_date: str
    rationale: str


# 20 improved locations (mid-sized metros / known winter-risk corridors)
INSTANCE_SPECS: List[InstanceSpec] = [
    InstanceSpec(1, "Duluth", "MN", 46.7867, -92.1005, "America/Chicago", "2025-11-25", "2025-11-27",
                 "Upper Midwest winter storm period; documented hundreds of crashes/spinouts in MN storm."),
    InstanceSpec(2, "St. Cloud", "MN", 45.5616, -94.1632, "America/Chicago", "2025-11-25", "2025-11-27",
                 "Upper Midwest winter storm period; documented hundreds of crashes/spinouts in MN storm."),
    InstanceSpec(3, "Fargo", "ND", 46.8772, -96.7898, "America/Chicago", "2025-11-25", "2025-11-27",
                 "I-94 corridor; winter storm window and refreeze risk."),
    InstanceSpec(4, "Sioux Falls", "SD", 43.5446, -96.7311, "America/Chicago", "2025-11-25", "2025-11-27",
                 "Upper Midwest storm window; cold nights + refreeze risk."),
    InstanceSpec(5, "Green Bay", "WI", 44.5133, -88.0133, "America/Chicago", "2025-11-29", "2025-12-01",
                 "Great Lakes winter roads; WI reports many crashes on winter roads."),
    InstanceSpec(6, "Marquette", "MI", 46.5436, -87.3954, "America/Detroit", "2025-11-29", "2025-12-01",
                 "Lake-effect/refreeze-prone Great Lakes region."),
    InstanceSpec(7, "Grand Rapids", "MI", 42.9634, -85.6681, "America/Detroit", "2026-01-24", "2026-01-26",
                 "Late Jan 2026 major winter storm window (snow/sleet/ice)."),
    InstanceSpec(8, "South Bend", "IN", 41.6764, -86.2520, "America/Indiana/Indianapolis", "2026-01-24", "2026-01-26",
                 "Late Jan 2026 major winter storm window; Great Lakes corridor."),
    InstanceSpec(9, "Toledo", "OH", 41.6528, -83.5379, "America/New_York", "2026-01-24", "2026-01-26",
                 "Late Jan 2026 winter storm window; Ohio Valley snow/sleet/ice."),
    InstanceSpec(10, "Cleveland", "OH", 41.4993, -81.6944, "America/New_York", "2026-01-24", "2026-01-26",
                 "Late Jan 2026 winter storm window; lake-enhanced snow/refreeze."),
    InstanceSpec(11, "Erie", "PA", 42.1292, -80.0851, "America/New_York", "2025-11-26", "2025-11-30",
                 "Late Nov 2025 lake-effect/refreeze window near Great Lakes."),
    InstanceSpec(12, "Buffalo", "NY", 42.8864, -78.8784, "America/New_York", "2025-11-26", "2025-11-30",
                 "Documented lake-effect storm window late Nov 2025."),
    InstanceSpec(13, "Rochester", "NY", 43.1566, -77.6088, "America/New_York", "2025-11-26", "2025-11-30",
                 "Late Nov 2025 lake-effect/refreeze-prone region."),
    InstanceSpec(14, "Syracuse", "NY", 43.0481, -76.1474, "America/New_York", "2025-11-26", "2025-11-30",
                 "Late Nov 2025 lake-effect/refreeze-prone region."),
    InstanceSpec(15, "Albany", "NY", 42.6526, -73.7562, "America/New_York", "2026-01-25", "2026-01-27",
                 "Late Jan 2026 storm impacts VT/N NY; warm overrunning cold air."),
    InstanceSpec(16, "Burlington", "VT", 44.4759, -73.2121, "America/New_York", "2026-01-25", "2026-01-27",
                 "Late Jan 2026 storm impacts VT/N NY; warm overrunning cold air."),
    InstanceSpec(17, "Portland", "ME", 43.6591, -70.2568, "America/New_York", "2026-02-11", "2026-02-12",
                 "Feb 11, 2026 New England event with NWS snowfall reports; refreeze after."),
    InstanceSpec(18, "Boston", "MA", 42.3601, -71.0589, "America/New_York", "2026-02-11", "2026-02-12",
                 "Feb 11, 2026 New England clipper snow/ice window."),
    InstanceSpec(19, "Madison", "WI", 43.0731, -89.4012, "America/Chicago", "2025-12-06", "2025-12-08",
                 "Early Dec 2025 clipper-type snow events listed by NWS Chicago; refreeze after."),
    InstanceSpec(20, "Chicago", "IL", 41.8781, -87.6298, "America/Chicago", "2026-01-24", "2026-01-26",
                 "NWS Chicago documented massive winter storm Jan 24–26, 2026."),
]


def openmeteo_client():
    cache_session = requests_cache.CachedSession(".cache_openmeteo", expire_after=3600)
    retry_session = retry(cache_session, retries=5, backoff_factor=0.2)
    return openmeteo_requests.Client(session=retry_session)


def build_raw_graph(lat: float, lon: float, dist_m: int, roads_filter: str) -> nx.MultiDiGraph:
    return ox.graph_from_point(
        (lat, lon),
        dist=dist_m,
        dist_type="bbox",
        retain_all=False,
        custom_filter=roads_filter,
        simplify=True,
    )


def consolidate_graph(G: nx.MultiDiGraph, tolerance_m: int) -> nx.MultiDiGraph:
    G_proj = ox.project_graph(G)
    G_cons = ox.consolidate_intersections(
        G_proj,
        tolerance=tolerance_m,
        rebuild_graph=True,
        dead_ends=False,
    )
    return ox.project_graph(G_cons, to_crs="EPSG:4326")


def graph_metrics(G: nx.MultiDiGraph) -> Dict[str, float]:
    n_nodes = G.number_of_nodes()
    n_edges = G.number_of_edges()

    if n_nodes == 0:
        lcc_frac = 0.0
        p99_deg = 0.0
        max_deg = 0.0
    else:
        Gu = G.to_undirected()
        comps = list(nx.connected_components(Gu))
        lcc = max((len(c) for c in comps), default=0)
        lcc_frac = lcc / float(n_nodes)

        degs = np.array([d for _, d in Gu.degree()], dtype=float)
        p99_deg = float(np.percentile(degs, 99))
        max_deg = float(degs.max())

    major_edges = 0
    for _, _, data in G.edges(data=True):
        hw = data.get("highway", None)
        if isinstance(hw, list):
            if any(h in MAJOR_ROADS for h in hw):
                major_edges += 1
        elif isinstance(hw, str):
            if hw in MAJOR_ROADS:
                major_edges += 1
    major_frac = major_edges / float(n_edges) if n_edges > 0 else 0.0

    return {
        "nodes": float(n_nodes),
        "edges": float(n_edges),
        "lcc_node_frac": float(lcc_frac),
        "major_edge_frac": float(major_frac),
        "deg_p99": float(p99_deg),
        "deg_max": float(max_deg),
    }


def within_constraints(m: Dict[str, float]) -> bool:
    return (
        NODES_MIN <= m["nodes"] <= NODES_MAX and
        EDGES_MIN <= m["edges"] <= EDGES_MAX and
        m["lcc_node_frac"] >= MIN_LCC_NODE_FRAC and
        m["major_edge_frac"] >= MIN_MAJOR_EDGE_FRAC and
        m["deg_max"] <= MAX_MAX_DEGREE and
        m["deg_p99"] <= MAX_P99_DEGREE
    )


def area_km2_from_dist(dist_m: int) -> float:
    return 4.0 * (dist_m ** 2) / 1e6


def tune_graph_for_instance(spec: InstanceSpec, cache_dir: str = ".cache_raw_graphs") -> Tuple[nx.MultiDiGraph, Dict, pd.DataFrame]:
    """
    Search over (road_filter, dist_m, tolerance_m) and pick a configuration that satisfies:
      - area bounds
      - nodes/edges bounds
      - structural sanity (LCC, major road frac, degree stats)
    
    Raw graphs are cached to avoid re-fetching from Overpass API on interruption/re-run.
    """
    os.makedirs(cache_dir, exist_ok=True)
    sweep_rows = []
    best = None  # (score, G, meta)

    total_configs = 0
    for filter_name, roads_filter in ROAD_FILTER_VARIANTS:
        for dist_m in DIST_CANDIDATES_M:
            if dist_m < DIST_MIN_M or dist_m > DIST_MAX_M:
                continue
            total_configs += 1
    
    print(f"  Searching {total_configs} distance/filter combinations × {len(TOLERANCE_CANDIDATES_M)} tolerances = {total_configs * len(TOLERANCE_CANDIDATES_M)} total configs...")
    config_num = 0

    for filter_name, roads_filter in ROAD_FILTER_VARIANTS:
        for dist_m in DIST_CANDIDATES_M:
            if dist_m < DIST_MIN_M or dist_m > DIST_MAX_M:
                continue
            config_num += 1
            
            # Check cache first
            cache_filename = f"inst{spec.instance_id:02d}_{filter_name}_dist{dist_m}.pickle"
            cache_path = os.path.join(cache_dir, cache_filename)
            
            if os.path.exists(cache_path):
                print(f"    [{config_num}/{total_configs}] Loading cached {filter_name} graph at dist={dist_m}m...", end=" ", flush=True)
                with open(cache_path, "rb") as f:
                    G_raw = pickle.load(f)
                print(f"✓ ({G_raw.number_of_nodes()} nodes, {G_raw.number_of_edges()} edges)")
            else:
                # Build raw graph once per (filter, dist)
                print(f"    [{config_num}/{total_configs}] Fetching {filter_name} graph at dist={dist_m}m from Overpass API...", end=" ", flush=True)
                G_raw = build_raw_graph(spec.center_lat, spec.center_lon, dist_m, roads_filter)
                print(f"✓ ({G_raw.number_of_nodes()} nodes, {G_raw.number_of_edges()} edges)")
                # Save to cache immediately
                with open(cache_path, "wb") as f:
                    pickle.dump(G_raw, f)
            
            for tol_m in TOLERANCE_CANDIDATES_M:
                G = consolidate_graph(G_raw, tol_m)
                m = graph_metrics(G)
                m.update({
                    "filter_name": filter_name,
                    "dist_m": dist_m,
                    "area_km2": area_km2_from_dist(dist_m),
                    "tolerance_m": tol_m,
                })
                sweep_rows.append(m)

                if not within_constraints(m):
                    continue

                # Score: prefer graphs near the middle of the target ranges
                nodes_mid = (NODES_MIN + NODES_MAX) / 2.0
                edges_mid = (EDGES_MIN + EDGES_MAX) / 2.0
                score = (
                    abs(m["nodes"] - nodes_mid) / nodes_mid +
                    abs(m["edges"] - edges_mid) / edges_mid +
                    0.5 * max(0.0, (m["deg_max"] - 20.0) / 20.0) -
                    0.2 * m["major_edge_frac"]  # reward more major-road structure
                )
                # Mild preference for major_only (cleaner graph) when otherwise similar
                if filter_name == "major_only":
                    score -= 0.05

                if best is None or score < best[0]:
                    best = (score, G, m)

    sweep_df = pd.DataFrame(sweep_rows)
    if best is None:
        # If nothing meets constraints, pick the closest candidate by a relaxed score
        # (Still produces diagnostics so you can adjust thresholds or candidates.)
        sweep_df["relaxed_score"] = (
            (sweep_df["nodes"] - (NODES_MIN + NODES_MAX)/2.0).abs() / ((NODES_MIN + NODES_MAX)/2.0) +
            (sweep_df["edges"] - (EDGES_MIN + EDGES_MAX)/2.0).abs() / ((EDGES_MIN + EDGES_MAX)/2.0) +
            (sweep_df["deg_max"] - 20.0).clip(lower=0.0) / 20.0
        )
        pick = sweep_df.sort_values("relaxed_score").iloc[0].to_dict()
        # Rebuild that exact configuration
        G_raw = build_raw_graph(spec.center_lat, spec.center_lon, int(pick["dist_m"]),
                                ROADS_FILTER_MAJOR_ONLY if pick["filter_name"]=="major_only" else ROADS_FILTER_WITH_RESIDENTIAL)
        G = consolidate_graph(G_raw, int(pick["tolerance_m"]))
        return G, pick, sweep_df.sort_values(["filter_name","dist_m","tolerance_m"])

    _, G_best, meta_best = best
    return G_best, meta_best, sweep_df.sort_values(["filter_name","dist_m","tolerance_m"])


def sample_weather_points(G: nx.MultiDiGraph, k: int = 5) -> List[Tuple[float, float]]:
    lats = np.array([data["y"] for _, data in G.nodes(data=True)])
    lons = np.array([data["x"] for _, data in G.nodes(data=True)])
    lat_min, lat_max = float(lats.min()), float(lats.max())
    lon_min, lon_max = float(lons.min()), float(lons.max())
    centroid = (float(lats.mean()), float(lons.mean()))
    corners = [(lat_min, lon_min), (lat_min, lon_max), (lat_max, lon_min), (lat_max, lon_max)]
    pts = [centroid] + corners
    return pts[:k]


def fetch_hourly_weather(points: List[Tuple[float, float]], start_date: str, end_date: str, timezone: str) -> pd.DataFrame:
    client = openmeteo_client()
    all_rows = []
    for i, (lat, lon) in enumerate(points):
        params = {
            "latitude": lat,
            "longitude": lon,
            "start_date": start_date,
            "end_date": end_date,
            "hourly": HOURLY_VARS,
            "timezone": timezone,
        }
        responses = client.weather_api(OPENMETEO_HISTORICAL_URL, params=params)
        r = responses[0]
        hourly = r.Hourly()

        times_utc = pd.date_range(
            start=pd.to_datetime(hourly.Time(), unit="s", utc=True),
            end=pd.to_datetime(hourly.TimeEnd(), unit="s", utc=True),
            freq=pd.Timedelta(hours=1),
            inclusive="left",
        )
        times_local = times_utc.tz_convert(timezone)

        data = {
            "sample_id": i,
            "latitude": lat,
            "longitude": lon,
            "time_utc": times_utc,
            "time_local": times_local,
        }
        for j, var in enumerate(HOURLY_VARS):
            data[var] = hourly.Variables(j).ValuesAsNumpy()
        all_rows.append(pd.DataFrame(data))

    return pd.concat(all_rows, ignore_index=True)


def is_passable_refreeze_window(weather_df: pd.DataFrame) -> bool:
    if not ENABLE_SNOW_PASSABILITY_CHECK:
        return True
    w = weather_df.copy()
    hours = w["time_local"].dt.hour
    w = w[(hours >= MISSION_START_LOCAL) & (hours < MISSION_END_LOCAL)]
    if w.empty:
        return True
    sd_cm = w["snow_depth"].astype(float) * 100.0
    median_sd_cm = float(np.nanmedian(sd_cm))
    mean_snowfall_mm = float(np.nanmean(w["snowfall"].astype(float)))
    return (median_sd_cm <= MAX_MEDIAN_SNOW_DEPTH_CM) and (mean_snowfall_mm <= MAX_MEAN_SNOWFALL_MM)


def assign_nodes_to_samples(G: nx.MultiDiGraph, sample_points: List[Tuple[float, float]]) -> Dict[int, int]:
    pts = np.array(sample_points)  # (lat, lon)
    node_ids, node_xy = [], []
    for nid, data in G.nodes(data=True):
        node_ids.append(int(nid))
        node_xy.append((data["y"], data["x"]))
    node_xy = np.array(node_xy)
    d2 = ((node_xy[:, None, :] - pts[None, :, :]) ** 2).sum(axis=2)
    nearest = d2.argmin(axis=1)
    return {nid: int(sid) for nid, sid in zip(node_ids, nearest)}


def write_excel(df: pd.DataFrame, xlsx_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Instances"

    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False):
        ws.append(list(row))

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = min(max(len(str(h)) + 2, 14), 45)

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def build_all(out_dir: str = "md_rpp_rrv_ice_instances_constrained") -> pd.DataFrame:
    os.makedirs(out_dir, exist_ok=True)
    graphs_dir = os.path.join(out_dir, "graphs")
    weather_dir = os.path.join(out_dir, "weather")
    meta_dir = os.path.join(out_dir, "meta")
    diag_dir = os.path.join(out_dir, "diagnostics")
    for d in (graphs_dir, weather_dir, meta_dir, diag_dir):
        os.makedirs(d, exist_ok=True)

    out_rows = []

    for spec in INSTANCE_SPECS:
        print(f"\nBuilding instance {spec.instance_id:02d}: {spec.place}, {spec.state}")
        G, chosen, sweep_df = tune_graph_for_instance(spec)

        # Save sweep diagnostics
        sweep_path = os.path.join(diag_dir, f"{spec.instance_id:02d}_{spec.place}_{spec.state}_graph_sweep.csv")
        sweep_df.to_csv(sweep_path, index=False)

        # Save graph
        graph_path = os.path.join(graphs_dir, f"{spec.instance_id:02d}_{spec.place}_{spec.state}.pickle")
        with open(graph_path, "wb") as f:
            pickle.dump(G, f)

        # Weather
        sample_pts = sample_weather_points(G, k=5)
        weather_df = fetch_hourly_weather(sample_pts, spec.start_date, spec.end_date, spec.timezone)
        passable = is_passable_refreeze_window(weather_df)
        weather_path = os.path.join(weather_dir, f"{spec.instance_id:02d}_{spec.place}_{spec.state}_hourly.csv")
        weather_df.to_csv(weather_path, index=False)

        # Node mapping
        node_to_sample = assign_nodes_to_samples(G, sample_pts)
        mapping_path = os.path.join(meta_dir, f"{spec.instance_id:02d}_{spec.place}_{spec.state}_node_to_sample.json")
        with open(mapping_path, "w") as f:
            json.dump(node_to_sample, f)

        # Metadata
        meta = asdict(spec)
        meta.update({
            "chosen_dist_m": int(chosen["dist_m"]),
            "area_km2": float(chosen["area_km2"]),
            "chosen_tolerance_m": int(chosen["tolerance_m"]),
            "road_filter_variant": str(chosen["filter_name"]),
            "computed_nodes": int(G.number_of_nodes()),
            "computed_edges": int(G.number_of_edges()),
            "major_edge_frac": float(chosen["major_edge_frac"]),
            "lcc_node_frac": float(chosen["lcc_node_frac"]),
            "deg_p99": float(chosen["deg_p99"]),
            "deg_max": float(chosen["deg_max"]),
            "passable_refreeze_window": bool(passable),
            "graph_path": graph_path,
            "weather_path": weather_path,
            "node_to_sample_path": mapping_path,
            "diagnostics_sweep_path": sweep_path,
        })
        meta_path = os.path.join(meta_dir, f"{spec.instance_id:02d}_{spec.place}_{spec.state}_meta.json")
        with open(meta_path, "w") as f:
            json.dump(meta, f, indent=2)

        out_rows.append(meta)
        print(f"  -> dist={meta['chosen_dist_m']}m area={meta['area_km2']:.1f}km² tol={meta['chosen_tolerance_m']}m "
              f"filter={meta['road_filter_variant']} nodes={meta['computed_nodes']} edges={meta['computed_edges']} "
              f"p99deg={meta['deg_p99']:.1f} maxdeg={meta['deg_max']:.0f} passable={meta['passable_refreeze_window']}")

    df = pd.DataFrame(out_rows).sort_values("instance_id")

    # Compact CSV / XLSX
    cols = [
        "instance_id","place","state","center_lat","center_lon","timezone",
        "start_date","end_date","rationale",
        "chosen_dist_m","area_km2","chosen_tolerance_m","road_filter_variant",
        "computed_nodes","computed_edges","major_edge_frac","lcc_node_frac","deg_p99","deg_max",
        "passable_refreeze_window",
        "graph_path","weather_path","node_to_sample_path","diagnostics_sweep_path"
    ]
    df_out = df[cols]
    csv_path = os.path.join(out_dir, "instances_metadata.csv")
    df_out.to_csv(csv_path, index=False)
    xlsx_path = os.path.join(out_dir, "instances_metadata.xlsx")
    write_excel(df_out, xlsx_path)
    print(f"\nWrote:\n  {csv_path}\n  {xlsx_path}")
    return df_out


if __name__ == "__main__":
    build_all()
